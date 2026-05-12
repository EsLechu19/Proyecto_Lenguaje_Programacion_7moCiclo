from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from productos.models import Producto, Categoria
from inventario.models import MovimientoStock

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter

@login_required(login_url='login')
def reportes(request):

    total_productos = Producto.objects.count()
    total_categorias = Categoria.objects.count()

    total_entradas = MovimientoStock.objects.filter(
        tipo='ENTRADA'
    ).count()

    total_salidas = MovimientoStock.objects.filter(
        tipo='SALIDA'
    ).count()

    productos_stock_bajo = Producto.objects.filter(
        stock__lte=5
    )

    ultimos_movimientos = MovimientoStock.objects.order_by(
        '-fecha'
    )[:10]

    return render(request, 'reportes/reportes.html', {
        'total_productos': total_productos,
        'total_categorias': total_categorias,
        'total_entradas': total_entradas,
        'total_salidas': total_salidas,
        'productos_stock_bajo': productos_stock_bajo,
        'ultimos_movimientos': ultimos_movimientos,
    })

@login_required(login_url='login')
def exportar_excel(request):
    workbook = Workbook()
    hoja = workbook.active
    hoja.title = "Reporte Inventario"

    hoja.append([
        "Producto",
        "Categoría",
        "Precio",
        "Stock actual",
        "Stock mínimo",
        "Estado"
    ])

    # ===== ESTILOS =====

    color_encabezado = PatternFill(
        start_color="1E3A8A",
        end_color="1E3A8A",
        fill_type="solid"
    )

    texto_blanco = Font(
        color="FFFFFF",
        bold=True
    )

    borde = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== ESTILOS DEL ENCABEZADO =====

    for celda in hoja[1]:
        celda.fill = color_encabezado
        celda.font = texto_blanco
        celda.alignment = Alignment(horizontal="center")
        celda.border = borde

    productos = Producto.objects.all()

    # ===== DATOS =====

    for producto in productos:

        estado = (
            "Stock bajo"
            if producto.stock <= producto.stock_minimo
            else "Normal"
        )

        hoja.append([
            producto.nombre,
            producto.categoria.nombre,
            producto.precio,
            producto.stock,
            producto.stock_minimo,
            estado
        ])

        fila = hoja.max_row

        for celda in hoja[fila]:
            celda.border = borde
            celda.alignment = Alignment(horizontal="center")

    # ===== AJUSTAR ANCHO DE COLUMNAS =====

    anchos = {
        'A': 25,
        'B': 20,
        'C': 15,
        'D': 15,
        'E': 15,
        'F': 20,
    }

    for columna, ancho in anchos.items():
        hoja.column_dimensions[columna].width = ancho

    # ===== RESPUESTA =====

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="reporte_inventario.xlsx"'
    )

    workbook.save(response)

    return response

@login_required(login_url='login')
def exportar_pdf(request):

    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = (
        'attachment; filename=\"reporte_inventario.pdf\"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=letter
    )

    elementos = []

    estilos = getSampleStyleSheet()

    titulo = Paragraph(
        "Reporte de Inventario - StockWise",
        estilos['Title']
    )

    elementos.append(titulo)
    elementos.append(Spacer(1, 20))

    productos = Producto.objects.all()

    datos = [[
        'Producto',
        'Categoría',
        'Precio',
        'Stock',
        'Estado'
    ]]

    for producto in productos:

        estado = (
            'Stock Bajo'
            if producto.stock <= producto.stock_minimo
            else 'Normal'
        )

        datos.append([
            producto.nombre,
            producto.categoria.nombre,
            f"S/ {producto.precio}",
            producto.stock,
            estado
        ])

    tabla = Table(datos)

    tabla.setStyle(TableStyle([

        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),

        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),

        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),

        ('GRID', (0, 0), (-1, -1), 1, colors.black),

        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),

    ]))

    elementos.append(tabla)

    doc.build(elementos)

    return response